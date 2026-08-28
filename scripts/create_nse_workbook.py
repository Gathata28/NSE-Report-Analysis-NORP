# Portable project root: set NORP_ROOT to override when invoked outside the repository.
from pathlib import Path as _Path
import os as _os
PROJECT_ROOT = _Path(_os.environ.get('NORP_ROOT', _Path(__file__).resolve().parents[1]))

from pathlib import Path
import csv, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
root=Path(str(PROJECT_ROOT))
wb=Workbook();ws=wb.active;ws.title='Report Index'
blue='1F4E78';light='D9EAF7';gray='F2F2F2';thin=Side(style='thin',color='D9E2F3')
def add_csv_sheet(name,path):
 sh=ws if name=='Report Index' else wb.create_sheet(name)
 rows=list(csv.reader(open(path,encoding='utf-8')))
 for row in rows:sh.append(row)
 sh.freeze_panes='A2';sh.auto_filter.ref=sh.dimensions
 if rows:
  ref=f'A1:{get_column_letter(len(rows[0]))}{len(rows)}'
  tab=Table(displayName=re.sub(r'[^A-Za-z0-9]','',name)[:20]+'Tbl',ref=ref)
  tab.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True,showColumnStripes=False)
  sh.add_table(tab)
  for cell in sh[1]:
   cell.font=Font(name='Arial',bold=True,color='FFFFFF');cell.fill=PatternFill('solid',fgColor=blue);cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
  for row in sh.iter_rows(min_row=2):
   for c in row:c.font=Font(name='Arial',size=10);c.alignment=Alignment(vertical='top',wrap_text=True)
  sh.row_dimensions[1].height=32
  for i,h in enumerate(rows[0],1):
   vals=[str(r[i-1]) for r in rows[1:min(len(rows),250)] if i-1<len(r)]
   width=min(65,max(12,len(str(h))+2,max((len(v) for v in vals),default=0)+2))
   sh.column_dimensions[get_column_letter(i)].width=width
  for row in sh.iter_rows(min_row=2):
   for c in row:
    if isinstance(c.value,str) and (c.value.startswith('http://') or c.value.startswith('https://')):
     c.hyperlink=c.value;c.font=Font(name='Arial',size=10,color='0563C1',underline='single')
 return sh
add_csv_sheet('Report Index',root/'nse_reports_normalized_validated.csv')
add_csv_sheet('Issuer Coverage',root/'current_issuer_coverage_log.csv')
add_csv_sheet('Validation Sample',root/'nse_link_validation_sample.csv')
add_csv_sheet('Historical Universe',root/'historical_nse_universe.csv')
add_csv_sheet('Gap Disclosures',root/'current_gap_disclosures.csv')
readme=wb.create_sheet('Readme',0)
readme_rows=[
 ['NSE Public Reports Archive — Readme',''],
 ['Collection as-of','2026-08-27'],
 ['Primary basis','Official issuer investor-relations/report pages and issuer-hosted PDFs or issuer-controlled asset/CDN endpoints.'],
 ['Fallback basis','NSE/CMA/archive sources are to be labeled separately; this tranche is predominantly issuer first-party.'],
 ['Scope taxonomy','Annual / full-year; Semi-annual / half-year; Quarterly; Periodic results material.'],
 ['Core report flag','Yes for annual reports and financial statements/results; No for presentations, press releases, commentaries, booklets, abridged or infograph materials.'],
 ['Validation interpretation','HTTP 200 or 206 with observed PDF content type is sample-validated. Official-page linkage without HTTP testing is retained as pending. 403 and request errors are not treated as functional downloads.'],
 ['Local files','Only files in the verified download manifest are populated with local path and SHA-256. The complete corpus was not downloaded.'],
 ['Fiscal period rule','Report year is derived from the document title/filename where available, not the web upload date. Blank or ambiguous periods remain blank rather than being fabricated.'],
 ['Current coverage','The coverage sheet compares the archive against the official 66-row current NSE universe. It is not a historical/delisted universe.'],
 ['Known limitations','Historical/delisted issuer enumeration remains incomplete; 1 current NSE issuer row still lacks an indexed financial-report record. Several fallback URLs remain pending or return slow/error responses.'],
 ['Responsible use','This is a source archive for analysis, not investment advice. Verify the linked document before relying on a record.'],
]
for r in readme_rows:readme.append(r)
readme.column_dimensions['A'].width=28;readme.column_dimensions['B'].width=115
for row in readme.iter_rows():
 for c in row:c.font=Font(name='Arial',size=10);c.alignment=Alignment(vertical='top',wrap_text=True)
for c in readme[1]:c.font=Font(name='Arial',bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor=blue)
readme.freeze_panes='A2'
wb.save(root/'nse_reports_archive_index.xlsx')
print('saved',root/'nse_reports_archive_index.xlsx','sheets',wb.sheetnames)
